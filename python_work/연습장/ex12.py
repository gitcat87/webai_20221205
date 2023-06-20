from openpyxl import Workbook,load_workbook
def Makestar(self):
        wb = Workbook()
        ws = wb.active
               
        for x in range(1,6):
            for y in range(1,x+1):
                ws.cell(row=x,column=y).value ="*"
                
        wb.save("MyStar.xlsx")
        wb.close() 

        lb = load_workbook("Mystar.xlsx")
        ls = lb.active

        for x in range(1,6):
            for y in range(1,x+1):
                print(ls.cell(row=x,column=y).value,end=" ")
                
        lb.close()