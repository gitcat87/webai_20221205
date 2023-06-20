from openpyxl import Workbook,load_workbook

class Star:

    def Makestar(self):
        wb = Workbook()
        ws = wb.active
               
        for x in range(1,6):
            for y in range(1,x+1):
                ws.cell(row=x,column=y).value ="*"
                
        wb.save("MyStar.xlsx")
        wb.close() 

    def Loadstar(self):
        lb = load_workbook("Mystar.xlsx")
        ls = lb.active

        for x in range(1,6):
            for y in range(1,x+1):

                print(ls['A1'].value,end='\n')
                print(ls['A2:B2'].value,end='\n')
                
        lb.close()
