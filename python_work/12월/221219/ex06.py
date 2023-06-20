from openpyxl import Workbook

wb = Workbook()
print(wb.sheetnames)

m1 = wb.create_sheet()
print(wb.sheetnames)

m1.title = "My시트"
m1.sheet_properties.tabColor='ff66ff'

ws1 = wb.create_sheet()
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
print(wb.sheetnames)

wb['Sheet1']["A1"] = "test"
wb['Sheet1']["A2"] = "test"
wb['Sheet1']["A3"] = "test"
wb['Sheet1']["A4"] = "test"

for i in range(1,10):
    wb['Sheet1'][f"A{i}"] = "test"

target = wb.copy_worksheet(wb["Sheet1"])
target.title("copy")

wb.save("sample.xlsx")
wb.close()