from openpyxl import Workbook, load_workbook
class MyWorkBook:

    
    def __init__(self):
        self.cnt = 1
        print("생성자 호출했음"+str(self.cnt))
        print(f"생성자 호출했음 {self.cnt}")
         #__init__는 초기화 함수

    def setCnt(self,cnt):
        self.cnt=cnt
        print("cnt바꿈")

    def doSave(self):
        print("저장했음.")

        wb = Workbook()
        ws = wb.active

        #ws['A1'] = "*"

        for c in range(1,self.cnt):#1에서 6전까지(5)
            for r in range(1,c+1):
                ws.cell(row = c, column=r,value='*')

        #cell(가로,세로,값) 또는 (가로,세로).값
        wb.save("star.xlsx")
        wb.close()

    def doLoad(self):
        print("불러왔음.")

        lb = load_workbook('star.xlsx')
        ws = lb.active

        for c in range(1,self.cnt):#1에서 6전까지(5)
            for r in range(1,c+1):
                print(ws.cell(row = c, column=r).value,end=" ")
            print()


        lb.close()

