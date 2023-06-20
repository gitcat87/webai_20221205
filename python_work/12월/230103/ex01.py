import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic


import lotto2
from openpyxl import Workbook,load_workbook

form_class = uic.loadUiType("12월/230103/lotto.ui")[0]

class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)       
               
        self.num1.setText("♥")        
        self.num2.setText("♥")
        self.num3.setText("♥")
        self.num4.setText("♥")
        self.num5.setText("♥")
        self.num6.setText("♥")
        self.num7.setText("★")
        self.numbtn.clicked.connect(self.doB)
    
        
    def doB(self):
        lw = load_workbook("lotto2.xlsx")
        ws = lw.active
    
        self.num1.setText(str(ws["A3"].value))
        self.num2.setText(str(ws["B3"].value))
        self.num3.setText(str(ws["C3"].value))
        self.num4.setText(str(ws["D3"].value))
        self.num5.setText(str(ws["E3"].value))
        self.num6.setText(str(ws["F3"].value))
        lw.save("lotto2.xlsx")
        lw.close()
        
if __name__ == "__main__" :
    app = QApplication(sys.argv) 
    myWindow = WindowClass() 
    myWindow.show()
    app.exec_()