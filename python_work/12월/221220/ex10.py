from openpyxl import Workbook,load_workbook
import random

class MyworkBook:

    def doSave(self):
        wb = Workbook() #워크북
        ws = wb.active #활성화 시트를 열고
        # 1행1열 10행10열까지 랜덤한 값을 넣어서
        for x in range(1,11):
            for y in range(1,11):
                ws.cell(row=y,column=x).value = random.randint(1,100)

        wb.save("random.xlsx") #save는 저장 파일
        wb.close()

    def doLoad(self):
        lb = load_workbook("random.xlsx")
        ws = lb.active
        for x in range(1,11):
            for y in range(1,11):
                print(ws.cell(row=y,column=x).value,end="")
         
        lb.close()

