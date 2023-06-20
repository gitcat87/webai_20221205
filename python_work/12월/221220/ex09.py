from openpyxl import Workbook,load_workbook
import random
# Workbook 덮어씌워지는거? load_workbook 불러오는거?
# from 라이브러리 import 모듈
# Workbook 파일 새로 생성
# load_workbook 파일 열기
# random 사용하면 랜덤 값 가져오기

'''
    A10 ~ J10
    1행1열 10행 10열까지 랜덤한 값을 넣어서
    random.xlsx 로 저장하고
    random.xlsx 불러와서
    1행열 ~ 10행 10열까지 출력하기
'''

wb = Workbook() #워크북
ws = wb.active #활성화 시트를 열고
# 1행1열 10행10열까지 랜덤한 값을 넣어서
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=y,column=x).value = random.randint(0,100)

wb.save("random.xlsx") #save는 저장 파일
wb.close()
#불러와서 출력하기
lb = load_workbook("random.xlsx")
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=y,column=x).value,end=" ")
    print()
lb.close()
